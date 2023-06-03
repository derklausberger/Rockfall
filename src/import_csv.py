import openpyxl
from datetime import datetime
from enum import Enum, auto

class FileType(Enum):
    AirHumidity = auto()
    AirTemp = auto()
    CrackmeterDistance = auto()
    Rainfall = auto()
    RockSurface = auto()
    RockTemp = auto()

#import crackmeterDistance data
def import_file(db, Measurements, filename, filetype: FileType):
    wb = openpyxl.load_workbook(filename)
    
    print(wb.sheetnames)
    #for sheet in wb.sheetnames:
    if True:
        sheet = "Brezno"
        ws = wb[sheet]
        for row in ws.iter_rows(2, 4400):#,ws.max_row):# 
            print(row)
            date = row[0].value
            if type(date) is str:
                date = datetime.strptime(date, "%d.%m.%Y")

            datetime_ = datetime.combine(
                date.date(),
                row[1].value
            )

            entry = db.session.execute(db.select(Measurements).filter_by(pilotCase=sheet, dateTime=datetime_)).scalar()

            insert = False

            if entry is None:
                insert = True
                entry = Measurements()
                
            if  (filetype == FileType.AirHumidity):
                entry.pilotCase = sheet
                entry.dateTime = datetime_
                entry.airHumidity = row[2].value
            elif  (filetype == FileType.AirTemp):
                entry.pilotCase = sheet
                entry.dateTime = datetime_
                entry.airTemp = row[2].value
            elif (filetype == FileType.CrackmeterDistance):
                entry.pilotCase = sheet
                entry.dateTime = datetime_
                entry.crackmeterDistance = row[2].value
            elif  (filetype == FileType.Rainfall):
                entry.pilotCase = sheet
                entry.dateTime = datetime_
                entry.rainfall = row[2].value
            elif  (filetype == FileType.RockSurface):
                entry.pilotCase = sheet
                entry.dateTime = datetime_
                entry.TRockSurface_A54C8C = row[2].value
                entry.TRockSurface_A54C8D = row[3].value
            elif  (filetype == FileType.RockTemp):
                entry.pilotCase = sheet
                entry.dateTime = datetime_
                entry.rockT25 = row[2].value
                entry.rockT50 = row[3].value
                entry.rockT75 = row[4].value
                entry.sensorT = row[5].value

            if insert == True:
                db.session.add(entry)

        db.session.commit()

#def import_db(db, Measurements):
    #import_file(db, Measurements, "Crackmeter_distance_data.xlsx", FileType.CrackmeterDistance)
    #import_file(db, Measurements, "Rock_Temp_data.xlsx", FileType.RockTemp)